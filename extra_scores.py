# gap, quant, std 전체 계산 모듈
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from decimal import Decimal, ROUND_HALF_UP

HEADER_FILL = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
HEADER_FONT = Font(bold=True)


# =========================
# 1. 데이터 로드 함수
# =========================

def get_close_data(filename):
    """
    '종가' 시트에서 날짜와 종가 시계열을 읽어온다.
    - dates: ['YYYYMMDD', ...]
    - stocks: [{'name': 종목명, 'code': 종목코드, 'prices': [가격 또는 None, ...]}, ...]
    """
    dates = []
    stocks = []
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb['종가']

        # 1행: 날짜 (3열부터 끝까지)
        raw_dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column + 1)]
        for d in raw_dates:
            if d is None:
                continue
            d_str = str(d)
            if len(d_str) == 8 and d_str.isdigit():
                if d_str not in dates:
                    dates.append(d_str)

        # 2행 이후: 종목명, 종목코드, 종가들
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row=row, column=1).value
            code = sheet.cell(row=row, column=2).value
            prices = [sheet.cell(row=row, column=col).value for col in range(3, sheet.max_column + 1)]
            prices_numeric = []
            for p in prices:
                try:
                    prices_numeric.append(float(p) if p not in (None, '') else None)
                except (ValueError, TypeError):
                    prices_numeric.append(None)
            stocks.append({'name': name, 'code': code, 'prices': prices_numeric})
    except Exception as e:
        print(f"⚠ 종가 시트 로딩 중 오류: {e}")
    return dates, stocks


def get_volume_data(filename):
    """
    '거래량' 시트에서 날짜와 거래량 시계열을 읽어온다.
    - dates: ['YYYYMMDD', ...]
    - stocks: [{'name': 종목명, 'code': 종목코드, 'volumes': [거래량 또는 None, ...]}, ...]
    """
    dates = []
    stocks = []
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb['거래량']

        raw_dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column + 1)]
        for d in raw_dates:
            if d is None:
                continue
            d_str = str(d)
            if len(d_str) == 8 and d_str.isdigit():
                if d_str not in dates:
                    dates.append(d_str)

        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row=row, column=1).value
            code = sheet.cell(row=row, column=2).value
            volumes = [sheet.cell(row=row, column=col).value for col in range(3, sheet.max_column + 1)]
            volumes_numeric = []
            for v in volumes:
                try:
                    volumes_numeric.append(int(v) if v not in (None, '') else None)
                except (ValueError, TypeError):
                    volumes_numeric.append(None)
            stocks.append({'name': name, 'code': code, 'volumes': volumes_numeric})
    except Exception as e:
        print(f"⚠ 거래량 시트 로딩 중 오류: {e}")
    return dates, stocks


def load_or_create_workbook(filename):
    try:
        return openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb


# =========================
# 2. GAP / QUANT / STD 계산식
# =========================

def calc_gap(prices_window20):
    """
    GAP 점수: (오늘 종가 / 최근 20일 평균) * 100
    (_gap.py의 calc_gap과 동일)
    """
    arr = [p for p in prices_window20 if p is not None]
    if len(arr) < 20:
        return None
    mean = np.mean(arr)
    if mean == 0:
        return 0
    val = 100 * (arr[-1] / mean)
    score = int(Decimal(str(val)).to_integral_value(rounding=ROUND_HALF_UP))
    return score


def calc_quant(volumes_window60):
    """
    QUANT 점수: ((현재 거래량 / 최근 60일 평균) * 100) / 2
    (_quant.py의 calc_quant와 동일)
    """
    arr = [v for v in volumes_window60 if v is not None]
    if len(arr) < 60:
        return None
    mean = np.mean(arr)
    if mean == 0:
        return 0
    val = ((arr[-1] / mean) * 100) / 2
    score = int(Decimal(str(val)).to_integral_value(rounding=ROUND_HALF_UP))
    return score


def calc_std_value(prices, idx, window_std=20, window_mean=20):
    """
    STD 값 계산 (_std.py의 calc_std_value 그대로):

    - idx 시점에서 20일 롤링 표준편차 σ_t 계산
    - 과거 window_mean(20)일 동안 각 시점의 20일 롤링 σ의 평균(avg_std) 계산
    - STD = (σ_t / avg_std - 1) * 100
    - 소수 둘째 자리까지 반올림
    """
    min_idx = window_std + window_mean - 2  # 예: 20 + 20 - 2 = 38
    if idx < min_idx:
        return None

    std_list = []
    for j in range(idx - window_mean + 1, idx + 1):  # j: idx-19 ~ idx
        start = j - window_std + 1
        end = j + 1
        if start < 0:
            return None

        window_prices = prices[start:end]

        if any(p is None for p in window_prices):
            return None

        arr = np.array(window_prices, dtype=float)
        sigma = float(np.std(arr, ddof=0))  # 모표준편차
        std_list.append(sigma)

    if not std_list:
        return None

    std_today = std_list[-1]
    avg_std = sum(std_list) / len(std_list)
    if avg_std == 0:
        return 0

    raw_val = (std_today / avg_std - 1) * 100
    val = float(Decimal(str(raw_val)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    return val


# =========================
# 3. 공통 시트 생성 유틸
# =========================

def get_existing_dates(sheet):
    dates = []
    for col in range(3, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        try:
            dates.append(int(val))
        except:
            continue
    return dates


def ensure_metric_sheet(wb, sheet_name, stocks):
    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        sheet.cell(row=1, column=1, value='종목명')
        sheet.cell(row=1, column=2, value='종목코드')
        sheet.cell(row=1, column=1).font = HEADER_FONT
        sheet.cell(row=1, column=2).font = HEADER_FONT
        sheet.cell(row=1, column=1).fill = HEADER_FILL
        sheet.cell(row=1, column=2).fill = HEADER_FILL

        code_to_row = {}
        for idx, stock in enumerate(stocks, start=2):
            code = str(stock['code'])
            sheet.cell(row=idx, column=1, value=stock['name'])
            sheet.cell(row=idx, column=2, value=code)
            code_to_row[code] = idx

        return sheet, [], code_to_row, []

    sheet = wb[sheet_name]
    existing_dates = get_existing_dates(sheet)

    sheet.cell(row=1, column=1, value='종목명').font = HEADER_FONT
    sheet.cell(row=1, column=2, value='종목코드').font = HEADER_FONT
    sheet.cell(row=1, column=1).fill = HEADER_FILL
    sheet.cell(row=1, column=2).fill = HEADER_FILL

    code_to_row = {}
    new_codes = []
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        code = sheet.cell(row=row, column=2).value
        if code is None:
            continue
        code_to_row[str(code)] = row

    for stock in stocks:
        code = str(stock['code'])
        if code not in code_to_row:
            max_row += 1
            sheet.cell(row=max_row, column=1, value=stock['name'])
            sheet.cell(row=max_row, column=2, value=code)
            code_to_row[code] = max_row
            new_codes.append(code)

    return sheet, existing_dates, code_to_row, new_codes


def write_header_cell(sheet, col_idx, date_value):
    cell = sheet.cell(row=1, column=col_idx, value=date_value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    sheet.column_dimensions[get_column_letter(col_idx)].width = 12


def fill_existing_for_new_codes(sheet, code_to_row, stock_map, existing_count, calc_func):
    if existing_count == 0:
        return
    for code, row_idx in code_to_row.items():
        stock = stock_map.get(code)
        if not stock or stock.get('_needs_backfill') is not True:
            continue
        series = stock['series']
        for idx_global in range(existing_count):
            col_idx = 3 + idx_global
            if sheet.cell(row=row_idx, column=col_idx).value not in (None, ""):
                continue
            val = calc_func(series, idx_global)
            if val is not None:
                sheet.cell(row=row_idx, column=col_idx, value=val)
        stock['_needs_backfill'] = False


def append_metric_columns(sheet, code_to_row, stock_map, existing_count, valid_dates, calc_func):
    for idx_global in range(existing_count, len(valid_dates)):
        date_val = valid_dates[idx_global]
        col_idx = 3 + idx_global
        write_header_cell(sheet, col_idx, date_val)

        for code, row_idx in code_to_row.items():
            stock = stock_map.get(code)
            if not stock:
                continue
            val = calc_func(stock['series'], idx_global)
            if val is not None:
                sheet.cell(row=row_idx, column=col_idx, value=val)


# =========================
# 4. GAP 시트 전체 재계산 저장
# =========================

def save_gap_sheet(filename, dates, stocks, window=20, sheet_name='gap'):
    if len(dates) < window:
        print(f"⚠ GAP: 날짜가 {window}일보다 적습니다. ({filename})")
        return

    valid_dates = dates[window - 1:]
    wb = load_or_create_workbook(filename)
    sheet, existing_dates, code_to_row, new_codes = ensure_metric_sheet(wb, sheet_name, stocks)

    stock_map = {}
    for stock in stocks:
        code = str(stock['code'])
        stock_map[code] = {'series': stock['prices']}
        if code in new_codes:
            stock_map[code]['_needs_backfill'] = True

    def calc_func(series, idx_global):
        end_idx = window - 1 + idx_global
        if end_idx >= len(series):
            return None
        window_prices = series[end_idx - window + 1:end_idx + 1]
        if None in window_prices:
            return None
        return calc_gap(window_prices)

    existing_count = len(existing_dates)
    fill_existing_for_new_codes(sheet, code_to_row, stock_map, existing_count, calc_func)

    if existing_count >= len(valid_dates):
        wb.save(filename)
        print(f"✅ GAP: 신규 날짜 없음 ({filename})")
        return

    append_metric_columns(sheet, code_to_row, stock_map, existing_count, valid_dates, calc_func)
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 12

    wb.save(filename)
    print(f"✅ GAP 업데이트 완료: {filename} (신규 {len(valid_dates) - existing_count}일)")


# =========================
# 5. QUANT 시트 전체 재계산 저장
# =========================

def save_quant_sheet(filename, dates, stocks, window=60, sheet_name='quant'):
    if len(dates) < window:
        print(f"⚠ QUANT: 날짜가 {window}일보다 적습니다. ({filename})")
        return

    valid_dates = dates[window - 1:]
    wb = load_or_create_workbook(filename)
    sheet, existing_dates, code_to_row, new_codes = ensure_metric_sheet(wb, sheet_name, stocks)

    stock_map = {}
    for stock in stocks:
        code = str(stock['code'])
        stock_map[code] = {'series': stock['volumes']}
        if code in new_codes:
            stock_map[code]['_needs_backfill'] = True

    def calc_func(series, idx_global):
        end_idx = window - 1 + idx_global
        if end_idx >= len(series):
            return None
        window_volumes = series[end_idx - window + 1:end_idx + 1]
        if None in window_volumes:
            return None
        return calc_quant(window_volumes)

    existing_count = len(existing_dates)
    fill_existing_for_new_codes(sheet, code_to_row, stock_map, existing_count, calc_func)

    if existing_count >= len(valid_dates):
        wb.save(filename)
        print(f"✅ QUANT: 신규 날짜 없음 ({filename})")
        return

    append_metric_columns(sheet, code_to_row, stock_map, existing_count, valid_dates, calc_func)
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 12

    wb.save(filename)
    print(f"✅ QUANT 업데이트 완료: {filename} (신규 {len(valid_dates) - existing_count}일)")


# =========================
# 6. STD 시트 전체 재계산 저장
# =========================

def save_std_sheet(filename, dates, stocks, sheet_name='std', window_std=20, window_mean=20):
    min_idx = window_std + window_mean - 2
    if len(dates) <= min_idx:
        print(f"⚠ STD: 데이터가 부족합니다. ({filename})")
        return

    valid_dates = dates[min_idx:]
    wb = load_or_create_workbook(filename)
    sheet, existing_dates, code_to_row, new_codes = ensure_metric_sheet(wb, sheet_name, stocks)

    stock_map = {}
    for stock in stocks:
        code = str(stock['code'])
        stock_map[code] = {'series': stock['prices']}
        if code in new_codes:
            stock_map[code]['_needs_backfill'] = True

    def calc_func(series, idx_global):
        i = min_idx + idx_global
        if i >= len(series):
            return None
        return calc_std_value(series, i, window_std=window_std, window_mean=window_mean)

    existing_count = len(existing_dates)
    fill_existing_for_new_codes(sheet, code_to_row, stock_map, existing_count, calc_func)

    if existing_count >= len(valid_dates):
        wb.save(filename)
        print(f"✅ STD: 신규 날짜 없음 ({filename})")
        return

    append_metric_columns(sheet, code_to_row, stock_map, existing_count, valid_dates, calc_func)
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 12

    wb.save(filename)
    print(f"✅ STD 업데이트 완료: {filename} (신규 {len(valid_dates) - existing_count}일)")


# =========================
# 7. 통합 실행 함수
# =========================

def run_extra_scores(filename):
    print(f"\n=== EXTRA SCORES 계산 시작: {filename} ===")

    # GAP, STD는 '종가' 시트 기반
    close_dates, close_stocks = get_close_data(filename)
    if close_dates and close_stocks:
        save_gap_sheet(filename, close_dates, close_stocks, window=20, sheet_name='gap')
        save_std_sheet(filename, close_dates, close_stocks, sheet_name='std', window_std=20, window_mean=20)
    else:
        print("⚠ 종가 데이터가 없어 GAP/STD 계산을 건너뜁니다.")

    # QUANT는 '거래량' 시트 기반
    vol_dates, vol_stocks = get_volume_data(filename)
    if vol_dates and vol_stocks:
        save_quant_sheet(filename, vol_dates, vol_stocks, window=60, sheet_name='quant')
    else:
        print("⚠ 거래량 데이터가 없어 QUANT 계산을 건너뜁니다.")

    print(f"=== EXTRA SCORES 계산 완료: {filename} ===\n")


def main():
    run_extra_scores("KR_Stocks_ETF.xlsx")


if __name__ == "__main__":
    main()
