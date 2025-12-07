import json
import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import time


# =========================
# 0. ì„¤ì •/ê³µí†µ ìœ í‹¸ í•¨ìˆ˜ë“¤
# =========================

def load_api_secrets(file_path='secrets.json'):
    """API í‚¤ì™€ ì‹œí¬ë¦¿ì„ íŒŒì¼ì—ì„œ ë¡œë“œ"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"ì—ëŸ¬: {file_path} íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        return None


def load_file_config(file_path='stock_file_map.json'):
    """ìì‚°ë³„ ì—‘ì…€ íŒŒì¼ ë§¤í•‘ ì •ë³´ ë¡œë“œ"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"ì—ëŸ¬: {file_path} íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        return None


def get_token(api_key, api_secret, domain):
    """í•œêµ­íˆ¬ìì¦ê¶Œ API í† í° ë°œê¸‰ ìš”ì²­"""
    url = f"{domain}/oauth2/tokenP"

    headers = {
        "content-type": "application/json",
        "appKey": api_key,
        "appSecret": api_secret
    }

    data = {
        "grant_type": "client_credentials",
        "appkey": api_key,
        "appsecret": api_secret
    }

    try:
        resp = requests.post(url, headers=headers, json=data)

        if resp.status_code != 200:
            print(f"âŒ í† í° ìš”ì²­ ì‹¤íŒ¨: HTTP {resp.status_code}")
            print(resp.text)
            return None

        token_data = resp.json()
        if not token_data or 'access_token' not in token_data:
            print("âŒ í† í° ì •ë³´ê°€ ì‘ë‹µì— ì—†ìŠµë‹ˆë‹¤")
            return None

        print("âœ… í† í° ë°œê¸‰ ì„±ê³µ!")
        return token_data

    except requests.exceptions.RequestException as e:
        print(f"âŒ í† í° ìš”ì²­ ì‹¤íŒ¨: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"ì„œë²„ ì‘ë‹µ: {e.response.text}")
        return None


# =========================
# 1. ì‹œì„¸ ì¡°íšŒ í•¨ìˆ˜ë“¤
# =========================

def fetch_stock_daily_history(access_token, domain, symbol, start_date, end_date,
                              app_key=None, app_secret=None):
    """
    êµ­ë‚´ ì£¼ì‹/ETF ê¸°ê°„ë³„ ì‹œì„¸ (ì¼ë³„)
    /uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice
    """
    endpoint = f"{domain}/uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice"

    params = {
        "FID_COND_MRKT_DIV_CODE": "J",    # ì£¼ì‹ ì‹œì¥ êµ¬ë¶„
        "FID_INPUT_ISCD": symbol,         # ì¢…ëª©ì½”ë“œ
        "FID_PERIOD_DIV_CODE": "D",       # ê¸°ê°„ êµ¬ë¶„ (ì¼)
        "FID_ORG_ADJ_PRC": "1",           # ìˆ˜ì •ì£¼ê°€ ì—¬ë¶€
        "FID_INPUT_DATE_1": start_date,   # ì¡°íšŒ ì‹œì‘ì¼
        "FID_INPUT_DATE_2": end_date,     # ì¡°íšŒ ì¢…ë£Œì¼
        "FID_COMP_ICD": symbol,
    }

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHKST03010100",     # ì£¼ì‹ ì¼ë³„ ì‹œì„¸
        "custtype": "P",
        "seq_no": "0",
        "locale": "ko_KR",
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"âŒ êµ­ë‚´ ì‹œì„¸ HTTP {resp.status_code} ì—ëŸ¬: {resp.text}")
            return None

        data = resp.json()
        if not data or 'output2' not in data or not data['output2']:
            # print("âŒ êµ­ë‚´ ì‹œì„¸ ë°ì´í„°ê°€ ë¹„ì–´ìˆìŠµë‹ˆë‹¤")
            return None

        daily_data = []
        for item in data['output2']:
            daily_data.append({
                'date': item.get('stck_bsop_date', ''),
                'open': int(item.get('stck_oprc', '0') or 0),
                'high': int(item.get('stck_hgpr', '0') or 0),
                'low': int(item.get('stck_lwpr', '0') or 0),
                'close': int(item.get('stck_clpr', '0') or 0),
                'volume': int(item.get('acml_vol', '0') or 0)
            })

        # ê³¼ê±° â†’ ìµœì‹  ì •ë ¬
        daily_data.sort(key=lambda x: x['date'])
        return daily_data

    except Exception as e:
        print(f"âŒ êµ­ë‚´ ì‹œì„¸ ì¡°íšŒ ì¤‘ ì—ëŸ¬: {str(e)}")
        return None


def fetch_overseas_daily_history(access_token, domain, market_code, symbol,
                                 start_date, end_date, app_key=None, app_secret=None):
    """
    í•´ì™¸ ì£¼ì‹/ETF ê¸°ê°„ë³„ ì‹œì„¸ (ì¼/ì£¼/ì›”)
    /uapi/overseas-price/v1/quotations/dailyprice
    - ì—¬ê¸°ì„œëŠ” ì¼ë´‰(GUBN=0), BYMD=end_date ê¸°ì¤€ìœ¼ë¡œ ìµœê·¼ 100ê°œ ë°›ì•„ì„œ
      date í•„í„°ë§(start_date~end_date)ë§Œ ì ìš©
    """
    endpoint = f"{domain}/uapi/overseas-price/v1/quotations/dailyprice"

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "HHDFS76240000",   # í•´ì™¸ ì£¼ì‹ ê¸°ê°„ë³„ ì‹œì„¸
        "custtype": "P",
    }

    params = {
        "AUTH": "",
        "EXCD": market_code,   # ì˜ˆ: "NAS"
        "SYMB": symbol,        # ì˜ˆ: "AAPL"
        "GUBN": "0",           # 0: ì¼, 1: ì£¼, 2: ì›”
        "BYMD": end_date,      # ê¸°ì¤€ì¼ (ì´ ë‚  í¬í•¨ ê³¼ê±° ë°©í–¥ ìµœëŒ€ 100ê°œ)
        "MODP": "0",           # 0: ì›ì£¼ê°€, 1: ìˆ˜ì •ì£¼ê°€
        # "KEYB": ""           # ì—°ì†ì¡°íšŒì‹œ ì‚¬ìš©, ì—¬ê¸°ì„œëŠ” ìƒëµ
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"âŒ í•´ì™¸ ì‹œì„¸ HTTP {resp.status_code} ì—ëŸ¬: {resp.text}")
            return None

        data = resp.json()
        rows = data.get("output2")
        if not rows:
            # print("âŒ í•´ì™¸ ì‹œì„¸ ë°ì´í„°ê°€ ë¹„ì–´ìˆìŠµë‹ˆë‹¤")
            return None

        daily_data = []
        for item in rows:
            d = item.get("xymd")   # ë‚ ì§œ (YYYYMMDD)
            if not d:
                continue

            # í•„ë“œëª…ì€ ì‹¤ì œ ì‘ë‹µì— ë”°ë¼ í•„ìš”ì‹œ í•œë²ˆ í™•ì¸
            daily_data.append({
                "date": d,
                "open": float(item.get("open", 0) or 0),
                "high": float(item.get("high", 0) or 0),
                "low": float(item.get("low", 0) or 0),
                "close": float(item.get("clos", 0) or 0),
                "volume": int(item.get("tvol", 0) or 0),
            })

        # ê³¼ê±° â†’ ìµœì‹ 
        daily_data.sort(key=lambda x: x["date"])

        # start_date~end_dateë¡œ í•„í„°ë§
        if start_date:
            daily_data = [d for d in daily_data if start_date <= d["date"] <= end_date]

        return daily_data

    except Exception as e:
        print(f"âŒ í•´ì™¸ ì‹œì„¸ ì¡°íšŒ ì¤‘ ì—ëŸ¬: {str(e)}")
        return None


# =========================
# 2. ì—‘ì…€ ê´€ë ¨ í•¨ìˆ˜ë“¤
# =========================

def load_stock_list(filename, market="KR"):
    """
    Excel íŒŒì¼ì—ì„œ ì¢…ëª© ëª©ë¡ì„ ì½ì–´ì˜µë‹ˆë‹¤.
    - market="KR" â†’ ì½”ë“œ 6ìë¦¬ zfill
    - market="US" â†’ ì½”ë“œ ê·¸ëŒ€ë¡œ ì‚¬ìš©
    """
    try:
        wb = openpyxl.load_workbook(filename)
        if "ì¢…ëª©" not in wb.sheetnames:
            print(f"\nâŒ Excel íŒŒì¼({filename})ì— 'ì¢…ëª©' ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤.")
            return None
        sheet = wb["ì¢…ëª©"]

        stocks = []
        for row in sheet.iter_rows(min_row=2):  # í—¤ë” ì œì™¸
            if row[0].value and row[1].value:
                raw_code = str(row[1].value).strip()
                if market == "KR":
                    code = raw_code.zfill(6)
                else:
                    code = raw_code

                stocks.append({
                    'name': row[0].value,
                    'code': code
                })

        print(f"\n[{filename}]ì—ì„œ ì½ì–´ì˜¨ ì¢…ëª© ëª©ë¡ ({market}):")
        for stock in stocks:
            print(f"  â€¢ {stock['name']} (ì½”ë“œ: {stock['code']})")

        return stocks

    except Exception as e:
        print(f"\nâŒ Excel íŒŒì¼ ì½ê¸° ì‹¤íŒ¨({filename}): {str(e)}")
        return None


def save_history_to_excel(data_list, filename, market="KR"):
    """
    ê° ì¢…ëª©ì˜ ì¼ë³„ OHLC ë°ì´í„°ë¥¼
    ì‹œê°€/ê³ ê°€/ì €ê°€/ì¢…ê°€/ê±°ë˜ëŸ‰ íƒ­ìœ¼ë¡œ ë‚˜ëˆ„ì–´ ì €ì¥.
    - í–‰: ì¢…ëª©
    - ì—´: ì¼ì
    market="KR"ì´ë©´ ì½”ë“œ 6ìë¦¬, "US"ë©´ ê·¸ëŒ€ë¡œ.
    """
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # ëª¨ë“  ë‚ ì§œ ìˆ˜ì§‘
    all_dates = set()
    for stock_data in data_list:
        if stock_data['history']:
            for daily in stock_data['history']:
                all_dates.add(daily['date'])

    sorted_dates = sorted(list(all_dates))
    if not sorted_dates:
        print("\nâŒ ì €ì¥í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    sheet_configs = [
        ('ì‹œê°€', 'open'),
        ('ê³ ê°€', 'high'),
        ('ì €ê°€', 'low'),
        ('ì¢…ê°€', 'close'),
        ('ê±°ë˜ëŸ‰', 'volume')
    ]

    for sheet_name, field_name in sheet_configs:
        # ê¸°ì¡´ ì‹œíŠ¸ ì—¬ë¶€
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            existing_dates = []
            for col in range(3, sheet.max_column + 1):
                val = sheet.cell(row=1, column=col).value
                try:
                    existing_dates.append(int(val))
                except Exception:
                    continue

            existing_data = {}
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                code = sheet.cell(row=row, column=2).value
                if not name or not code:
                    continue
                code_str = str(code).strip()
                if market == "KR":
                    code_key = code_str.zfill(6)
                else:
                    code_key = code_str

                values = {}
                for col_idx, date_int in enumerate(existing_dates, 3):
                    values[str(date_int)] = sheet.cell(row=row, column=col_idx).value
                existing_data[code_key] = {'name': name, 'values': values}
        else:
            sheet = wb.create_sheet(sheet_name)
            existing_dates = []
            existing_data = {}

        merged_dates = set(existing_dates)
        for stock_data in data_list:
            if stock_data['history']:
                for daily in stock_data['history']:
                    try:
                        merged_dates.add(int(daily['date']))
                    except Exception:
                        continue

        sorted_dates_all = sorted(list(merged_dates))

        # í—¤ë”
        sheet.cell(row=1, column=1, value='ì¢…ëª©ëª…')
        sheet.cell(row=1, column=2, value='ì¢…ëª©ì½”ë“œ')
        for col_idx, date_int in enumerate(sorted_dates_all, 3):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = date_int
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        for col_idx in (1, 2):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # ì¢…ëª© ì½”ë“œ ì „ì²´ ì§‘í•©
        all_codes = set(existing_data.keys())
        for stock_data in data_list:
            all_codes.add(stock_data['code'])

        for row_idx, code in enumerate(sorted(all_codes), start=2):
            # ì´ë¦„
            if code in existing_data:
                name = existing_data[code]['name']
            else:
                # data_listì—ì„œ ì°¾ê¸°
                name = next((s['name'] for s in data_list if s['code'] == code), code)

            sheet.cell(row=row_idx, column=1, value=name)
            sheet.cell(row=row_idx, column=2, value=code)

            # ê¸°ì¡´ ê°’
            values = existing_data.get(code, {}).get('values', {})

            # ì‹ ê·œ ê°’
            new_values = {}
            stock_hist = next((s for s in data_list if s['code'] == code), None)
            if stock_hist and stock_hist['history']:
                for daily in stock_hist['history']:
                    try:
                        new_values[str(int(daily['date']))] = daily[field_name]
                    except Exception:
                        continue

            # ë‚ ì§œë³„ë¡œ ê°’ ì…ë ¥
            for col_idx, date_int in enumerate(sorted_dates_all, 3):
                key = str(date_int)
                val = new_values.get(key, values.get(key, ''))
                sheet.cell(row=row_idx, column=col_idx, value=val)

        # ì—´ ë„ˆë¹„
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 14
        for col_idx in range(3, len(sorted_dates_all) + 3):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 12

    wb.save(filename)
    print(f"\nâœ… ì—‘ì…€ íŒŒì¼ ì €ì¥ ì™„ë£Œ: {filename}")


def get_latest_date_from_sheet(filename, sheet_name):
    """ì§€ì • ì‹œíŠ¸(ì¢…ê°€/ê±°ë˜ëŸ‰ ë“±)ì—ì„œ ê°€ì¥ ìµœì‹  ë‚ ì§œë¥¼ 'YYYYMMDD' ë¬¸ìì—´ë¡œ ë°˜í™˜"""
    try:
        wb = openpyxl.load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return None
        sheet = wb[sheet_name]

        dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column + 1)]
        dates_dt = []
        for d in dates:
            try:
                dates_dt.append(datetime.strptime(str(d), '%Y%m%d'))
            except Exception:
                pass
        if not dates_dt:
            return None
        latest = max(dates_dt)
        return latest.strftime('%Y%m%d')
    except Exception as e:
        print(f"âŒ ë‚ ì§œ ì¶”ì¶œ ì—ëŸ¬({filename}/{sheet_name}): {e}")
        return None


# =========================
# 3. ì§€ìˆ˜(ì½”ìŠ¤í”¼/ì½”ìŠ¤ë‹¥) ì‹œíŠ¸
# =========================

def fetch_index_history(access_token, domain, index_code, app_key, app_secret,
                        start_date, end_date):
    """
    ì—…ì¢…ì§€ìˆ˜ ê¸°ê°„ë³„ ì‹œì„¸ ì¡°íšŒ (ì¼ë³„)
    - index_code: 0001(KOSPI), 1001(KOSDAQ), 2001(KOSPI200)
    """
    endpoint = f"{domain}/uapi/domestic-stock/v1/quotations/inquire-daily-indexchartprice"

    params = {
        "fid_cond_mrkt_div_code": "U",
        "fid_input_iscd": index_code,
        "fid_input_date_1": start_date,
        "fid_input_date_2": end_date,
        "fid_period_div_code": "D",
    }

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHKUP03500100",
        "custtype": "P",
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"âŒ ì—…ì¢…ì§€ìˆ˜ HTTP {resp.status_code} ì˜¤ë¥˜ ({index_code})")
            print(resp.text)
            return None

        data = resp.json()
        rows = data.get("output2")

        if not rows:
            print(f"âŒ ì—…ì¢…ì§€ìˆ˜ ë°ì´í„° ì—†ìŒ ({index_code})")
            return None

        history = []
        for row in rows:
            history.append({
                "date": row.get("stck_bsop_date", ""),
                "index_value": row.get("bstp_nmix_prpr"),
                "open": row.get("bstp_nmix_oprc"),
                "high": row.get("bstp_nmix_hgpr"),
                "low": row.get("bstp_nmix_lwpr"),
            })

        history.sort(key=lambda x: x["date"])
        return history

    except Exception as e:
        print(f"âŒ ì—…ì¢…ì§€ìˆ˜ ì¡°íšŒ ì¤‘ ì—ëŸ¬ ({index_code}): {e}")
        return None


def update_index_sheet(access_token, domain, app_key, app_secret,
                       filename="KR_Stocks_Individual.xlsx"):
    """
    íŒŒì¼ì˜ 'ì§€ìˆ˜' ì‹œíŠ¸ ì—…ë°ì´íŠ¸
    - ì—†ìœ¼ë©´: ìµœê·¼ 100ì¼ì¹˜ KOSPI/KOSDAQ/KOSPI200 ìƒì„±
    - ìˆìœ¼ë©´: ë§ˆì§€ë§‰ ë‚ ì§œ ì´í›„ ~ ì˜¤ëŠ˜ê¹Œì§€ ì¶”ê°€
    """
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    indices = [
        ("KOSPI", "0001"),
        ("KOSDAQ", "1001"),
        ("KOSPI200", "2001"),
    ]

    today = datetime.now()
    today_str = today.strftime('%Y%m%d')

    # A. ì‹œíŠ¸ ì—†ëŠ” ê²½ìš°
    if 'ì§€ìˆ˜' not in wb.sheetnames:
        sheet = wb.create_sheet('ì§€ìˆ˜')

        end_date = today_str
        start_date = (today - timedelta(days=100)).strftime('%Y%m%d')
        print(f"\nğŸ“ˆ [ì§€ìˆ˜] ìµœì´ˆ ìƒì„±: {start_date} ~ {end_date}")

        index_data = {}
        all_dates = set()

        for name, code in indices:
            print(f"  â–¶ {name} ({code}) ì¡°íšŒ ì¤‘...")
            history = fetch_index_history(
                access_token, domain, code, app_key, app_secret,
                start_date, end_date
            )
            if not history:
                print(f"    â€¢ {name} ë°ì´í„° ì—†ìŒ")
                continue

            values = {}
            for h in history:
                d = h["date"]
                v = h["index_value"]
                if not d or v is None:
                    continue
                values[d] = float(v)
                all_dates.add(d)

            index_data[code] = {
                "name": name,
                "code": code,
                "values": values
            }
            print(f"    â€¢ {len(values)}ì¼ì¹˜ ë°ì´í„° í™•ë³´")

            time.sleep(0.5)

        if not index_data or not all_dates:
            print("\nâŒ ì§€ìˆ˜ ë°ì´í„°ê°€ ì—†ì–´ ì €ì¥í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.")
            wb.save(filename)
            return

        sorted_dates = sorted(all_dates)

        # í—¤ë”
        sheet.cell(row=1, column=1, value='ì—…ì¢…ëª…')
        sheet.cell(row=1, column=2, value='ì—…ì¢…ì½”ë“œ')
        sheet.cell(row=1, column=1).font = Font(bold=True)
        sheet.cell(row=1, column=2).font = Font(bold=True)
        sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        for col_idx, date_str in enumerate(sorted_dates, 3):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = date_str
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # ë°ì´í„°
        for row_idx, code in enumerate(sorted(index_data.keys()), start=2):
            info = index_data[code]
            sheet.cell(row=row_idx, column=1, value=info["name"])
            sheet.cell(row=row_idx, column=2, value=info["code"])

            values = info["values"]
            for col_idx, date_str in enumerate(sorted_dates, 3):
                val = values.get(date_str, "")
                sheet.cell(row=row_idx, column=col_idx, value=val)

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 12
        for col_idx in range(3, len(sorted_dates) + 3):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 12

        wb.save(filename)
        print(f"\nâœ… 'ì§€ìˆ˜' ì‹œíŠ¸ ìµœì´ˆ ìƒì„± ì™„ë£Œ: {filename}")
        return

    # B. ì‹œíŠ¸ ìˆëŠ” ê²½ìš° â†’ ì¶”ê°€
    sheet = wb['ì§€ìˆ˜']
    print("\nğŸ“ˆ [ì§€ìˆ˜] ê¸°ì¡´ ì‹œíŠ¸ ì—…ë°ì´íŠ¸ ì‹œì‘")

    existing_dates = []
    for col in range(3, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val:
            existing_dates.append(str(val))

    existing_data = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        code = sheet.cell(row=row, column=2).value
        if not code:
            continue
        code_str = str(code).strip()
        values = {}
        for idx, date_str in enumerate(existing_dates, 3):
            values[date_str] = sheet.cell(row=row, column=idx).value
        existing_data[code_str] = {"name": name, "values": values}

    latest = get_latest_date_from_sheet(filename, "ì§€ìˆ˜")
    if latest:
        start_dt = datetime.strptime(latest, "%Y%m%d") + timedelta(days=1)
        start_date = start_dt.strftime("%Y%m%d")
        print(f"  â€¢ ë§ˆì§€ë§‰ ë‚ ì§œ: {latest} â†’ ì¶”ê°€ ì¡°íšŒ ì‹œì‘ì¼: {start_date}")
    else:
        start_date = (today - timedelta(days=100)).strftime("%Y%m%d")
        print(f"  â€¢ ê¸°ì¡´ ë‚ ì§œ ì—†ìŒ â†’ {start_date} ~ {today_str} ì¬ì¡°íšŒ")

    end_date = today_str
    if datetime.strptime(start_date, "%Y%m%d") > datetime.strptime(end_date, "%Y%m%d"):
        print("  â€¢ ì¶”ê°€í•  ì§€ìˆ˜ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤. (ì´ë¯¸ ìµœì‹ )")
        return

    new_index_data = {}
    all_dates = set(existing_dates)

    for name, code in indices:
        print(f"  â–¶ {name} ({code}) ì‹ ê·œ ì¡°íšŒ: {start_date} ~ {end_date}")
        history = fetch_index_history(
            access_token, domain, code, app_key, app_secret,
            start_date, end_date
        )
        if not history:
            print(f"    â€¢ {name} ì¶”ê°€ ë°ì´í„° ì—†ìŒ")
            continue

        values = {}
        for h in history:
            d = h["date"]
            v = h["index_value"]
            if not d or v is None:
                continue
            values[d] = float(v)
            all_dates.add(d)

        new_index_data[code] = {
            "name": name,
            "code": code,
            "values": values
        }
        print(f"    â€¢ {len(values)}ì¼ì¹˜ ì‹ ê·œ ë°ì´í„° í™•ë³´")

        time.sleep(0.5)

    if not new_index_data:
        print("  â€¢ ì‹ ê·œ ì§€ìˆ˜ ë°ì´í„°ê°€ ì—†ì–´ ì—…ë°ì´íŠ¸í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.")
        return

    merged_dates = sorted(all_dates)

    # í—¤ë”
    sheet.cell(row=1, column=1, value='ì—…ì¢…ëª…')
    sheet.cell(row=1, column=2, value='ì—…ì¢…ì½”ë“œ')
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    for col_idx, date_str in enumerate(merged_dates, 3):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = date_str
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    all_codes = set(existing_data.keys()) | set([code for _, code in indices])

    for row_idx, code in enumerate(sorted(all_codes), start=2):
        if code in existing_data:
            name = existing_data[code]["name"]
        else:
            name = next((n for (n, c) in indices if c == code), code)

        sheet.cell(row=row_idx, column=1, value=name)
        sheet.cell(row=row_idx, column=2, value=code)

        old_values = existing_data.get(code, {}).get("values", {})
        new_values = new_index_data.get(code, {}).get("values", {})

        for col_idx, date_str in enumerate(merged_dates, 3):
            val = new_values.get(date_str, old_values.get(date_str, ""))
            sheet.cell(row=row_idx, column=col_idx, value=val)

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 12
    for col_idx in range(3, len(merged_dates) + 3):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 12

    wb.save(filename)
    print(f"\nâœ… 'ì§€ìˆ˜' ì‹œíŠ¸ ì—…ë°ì´íŠ¸ ì™„ë£Œ: {filename}")


# =========================
# 4. ì¹´í…Œê³ ë¦¬ë³„ ì²˜ë¦¬ ë˜í¼
# =========================

def fetch_kr_wrapper(access_token, domain, symbol, start_date, end_date,
                     app_key, app_secret):
    """êµ­ë‚´(ê°œë³„+ETF) ëª¨ë‘ ë™ì¼ API ì‚¬ìš©"""
    return fetch_stock_daily_history(
        access_token, domain, symbol, start_date, end_date,
        app_key=app_key, app_secret=app_secret
    )


def fetch_us_wrapper(access_token, domain, symbol, start_date, end_date,
                     app_key, app_secret):
    """
    ë¯¸êµ­(ê°œë³„+ETF) ëª¨ë‘ í•´ì™¸ ê¸°ê°„ë³„ ì‹œì„¸ API ì‚¬ìš©
    - ì—¬ê¸°ì„œëŠ” ì¼ë‹¨ NASDAQ("NAS") ê¸°ì¤€ìœ¼ë¡œ í˜¸ì¶œ
      (ë‚˜ì¤‘ì— í•„ìš”í•˜ë©´ ì—‘ì…€ì— EXCD ì»¬ëŸ¼ ì¶”ê°€í•´ì„œ í™•ì¥)
    """
    return fetch_overseas_daily_history(
        access_token, domain, market_code="NAS", symbol=symbol,
        start_date=start_date, end_date=end_date,
        app_key=app_key, app_secret=app_secret
    )


def process_one_file(excel_filename, fetch_func, app_key, app_secret, domain,
                     access_token, market="KR", update_index=False):
    """
    ì—‘ì…€ íŒŒì¼ 1ê°œ ì²˜ë¦¬:
    - ì¢…ëª© ëª©ë¡ ë¡œë“œ
    - ì¢…ê°€/ê±°ë˜ëŸ‰ ì‹œíŠ¸ ê¸°ì¤€ìœ¼ë¡œ ë‚ ì§œ ë²”ìœ„ ê²°ì •
    - fetch_funcìœ¼ë¡œ ê° ì¢…ëª© íˆìŠ¤í† ë¦¬ ê°€ì ¸ì˜¤ê¸°
    - ì‹œê°€/ê³ ê°€/ì €ê°€/ì¢…ê°€/ê±°ë˜ëŸ‰ ì‹œíŠ¸ ì €ì¥
    - í•„ìš” ì‹œ ì§€ìˆ˜ ì‹œíŠ¸ ì—…ë°ì´íŠ¸
    """
    stocks = load_stock_list(excel_filename, market=market)
    if not stocks:
        return

    latest_close = get_latest_date_from_sheet(excel_filename, "ì¢…ê°€")
    latest_amount = get_latest_date_from_sheet(excel_filename, "ê±°ë˜ëŸ‰")

    today = datetime.now()
    today_str = today.strftime('%Y%m%d')

    if latest_close and latest_amount:
        latest_str = max(latest_close, latest_amount)
        start_dt = datetime.strptime(latest_str, '%Y%m%d') + timedelta(days=1)
        start_date = start_dt.strftime('%Y%m%d')
        print(f"\nğŸ“… [{excel_filename}] ì¶”ê°€ ì¡°íšŒ: {start_date} ~ {today_str}")
    else:
        end_dt = today
        start_date = (end_dt - timedelta(days=100)).strftime('%Y%m%d')
        print(f"\nğŸ“… [{excel_filename}] ì „ì²´ ì¡°íšŒ(ìµœê·¼ 100ì¼): {start_date} ~ {today_str}")

    end_date = today_str

    data_list = []
    print(f"\nì´ {len(stocks)}ê°œ ì¢…ëª©ì— ëŒ€í•´ ì¡°íšŒí•©ë‹ˆë‹¤... ({excel_filename})")
    for i, stock in enumerate(stocks, start=1):
        print(f"  [{i}/{len(stocks)}] {stock['name']}({stock['code']}) ...", end='')

        history = fetch_func(
            access_token, domain,
            stock['code'], start_date, end_date,
            app_key, app_secret
        )

        if history:
            print(f"ì„±ê³µ ({len(history)}ì¼)")
            data_list.append({
                "name": stock['name'],
                "code": stock['code'],
                "history": history,
            })
        else:
            print("ì‹¤íŒ¨ ë˜ëŠ” ì¶”ê°€ ë°ì´í„° ì—†ìŒ")

        time.sleep(1)

    if data_list:
        save_history_to_excel(data_list, filename=excel_filename, market=market)
        if update_index and market == "KR":
            update_index_sheet(
                access_token=access_token,
                domain=domain,
                app_key=app_key,
                app_secret=app_secret,
                filename=excel_filename
            )
    else:
        print(f"\nâŒ [{excel_filename}] ì €ì¥í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")


# =========================
# 5. main
# =========================

def main():
    print(f"\n=== í•œêµ­íˆ¬ìì¦ê¶Œ API ì£¼ì‹/ETF ì‹œì„¸ íˆìŠ¤í† ë¦¬ ì¡°íšŒ ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ===")

    secrets = load_api_secrets()
    if not secrets:
        return

    app_key = secrets.get('api_key')
    app_secret = secrets.get('api_secret')
    domain = secrets.get('domain', 'https://openapi.koreainvestment.com:9443')

    file_config = load_file_config('stock_file_map.json')
    if not file_config:
        return

    print("\nğŸ”„ í† í° ë°œê¸‰ ìš”ì²­ ì¤‘...")
    token_data = get_token(app_key, app_secret, domain)
    if not token_data:
        print("\nâŒ í† í° ë°œê¸‰ ì‹¤íŒ¨")
        return
    access_token = token_data['access_token']

    # ì¹´í…Œê³ ë¦¬ë³„ ì„¤ì •
    # - market: ì½”ë“œ ì²˜ë¦¬(KR: zfill, US: ê·¸ëŒ€ë¡œ)
    # - fetch_func: ì–´ë–¤ API í˜¸ì¶œí• ì§€
    # - update_index: ì§€ìˆ˜ ì‹œíŠ¸ ìƒì„±/ì—…ë°ì´íŠ¸ ì—¬ë¶€
    category_settings = {
        "KR_Stocks_Individual": {"market": "KR", "fetch_func": fetch_kr_wrapper, "update_index": True},
        "KR_Stocks_ETF":        {"market": "KR", "fetch_func": fetch_kr_wrapper, "update_index": False},
        "US_Stocks_Individual": {"market": "US", "fetch_func": fetch_us_wrapper, "update_index": False},
        "US_Stocks_ETF":        {"market": "US", "fetch_func": fetch_us_wrapper, "update_index": False},
    }

    for category_name, excel_filename in file_config.items():
        print("\n=======================================")
        print(f"ğŸ“‚ ì¹´í…Œê³ ë¦¬: {category_name}")
        print(f"ğŸ“Š íŒŒì¼: {excel_filename}")
        print("=======================================")

        cfg = category_settings.get(category_name)
        if not cfg:
            print(f"âš ï¸ {category_name} ì— ëŒ€í•œ ì„¤ì •ì´ ì—†ìŠµë‹ˆë‹¤. ê±´ë„ˆëœ€.")
            continue

        process_one_file(
            excel_filename=excel_filename,
            fetch_func=cfg["fetch_func"],
            app_key=app_key,
            app_secret=app_secret,
            domain=domain,
            access_token=access_token,
            market=cfg["market"],
            update_index=cfg["update_index"]
        )


if __name__ == "__main__":
    main()
