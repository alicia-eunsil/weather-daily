# totalS, totalZ 통합모듈

import openpyxl
from openpyxl.styles import Font, PatternFill
from decimal import Decimal, ROUND_HALF_UP
import numpy as np


# =========================
# 1. S 점수 (원본 totalS.py 동일)
# =========================
def calc_s(prices, window: int):
    arr = [p for p in prices if p is not None]
    if len(arr) < window:
        return None

    arr = arr[-window:]  # 최근 window일

    min_val = min(arr)
    max_val = max(arr)
    if max_val == min_val:
        return 0

    val = 100 * ((arr[-1] - min_val) / (max_val - min_val))

    score = int(Decimal(str(val)).to_integral_value(rounding=ROUND_HALF_UP))
    return score


# =========================
# 2. Z 점수 (원본 totalZ.py 동일)
# =========================
def calc_z(prices, window: int):
    arr = [p for p in prices if p is not None]
    if len(arr) < window:
        return None

    arr = arr[-window:]  # 최근 window일

    mean = float(np.mean(arr))
    std = float(np.std(arr, ddof=1))  # 표본 표준편차

    if std == 0:
        return 0

    z = (arr[-1] - mean) / std
    val = 50 * z  

    score = int(Decimal(str(val)).to_integral_value(rounding=ROUND_HALF_UP))
    return score


# =========================
# 3. 종가 시트를 읽어서 dates, stocks 반환
# =========================
def get_close_data(filename: str):
    wb = openpyxl.load_workbook(filename, data_only=True)

    if "종가" not in wb.sheetnames:
        raise ValueError(f"'{filename}' 파일에 '종가' 시트가 없습니다.")

    sheet = wb["종가"]

    # 날짜 가져오기
    dates = []
    for col in range(3, sheet.max_column + 1):
        v = sheet.cell(row=1, column=col).value
        if v is None:
            continue
        try:
            dates.append(int(v))
        except:
            continue

    # 종목 데이터
    stocks = []
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        code = sheet.cell(row=row, column=2).value
        if not name or not code:
            continue

        prices = []
        for col in range(3, sheet.max_column + 1):
            v = sheet.cell(row=row, column=col).value
            if v is None or v == "":
                prices.append(None)
            else:
                try:
                    prices.append(float(v))
                except:
                    prices.append(None)

        stocks.append({
            "name": name,
            "code": str(code),
            "prices": prices,
        })

    return dates, stocks


# =========================
# 4. 기존 시트 삭제 후 새로 생성
# =========================
def get_existing_dates(sheet):
    dates = []
    for col in range(3, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        try:
            dates.append(int(val))
        except:
            continue
    return dates


def ensure_score_sheet(wb, sheet_name, stocks):
    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        sheet.cell(row=1, column=1, value="종목명")
        sheet.cell(row=1, column=2, value="종목코드")
        sheet.cell(row=1, column=1).font = Font(bold=True)
        sheet.cell(row=1, column=2).font = Font(bold=True)
        sheet.cell(row=1, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        sheet.cell(row=1, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        code_to_row = {}
        for idx, stock in enumerate(stocks, start=2):
            code = str(stock["code"])
            sheet.cell(row=idx, column=1, value=stock["name"])
            sheet.cell(row=idx, column=2, value=code)
            code_to_row[code] = idx

        return sheet, [], code_to_row, []

    sheet = wb[sheet_name]
    existing_dates = get_existing_dates(sheet)

    # 보정: 1,2열 헤더 강조
    sheet.cell(row=1, column=1, value="종목명").font = Font(bold=True)
    sheet.cell(row=1, column=2, value="종목코드").font = Font(bold=True)
    sheet.cell(row=1, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    sheet.cell(row=1, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    code_to_row = {}
    new_codes = []
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        code = sheet.cell(row=row, column=2).value
        if code is None:
            continue
        code_to_row[str(code)] = row

    for stock in stocks:
        code = str(stock["code"])
        if code not in code_to_row:
            max_row += 1
            sheet.cell(row=max_row, column=1, value=stock["name"])
            sheet.cell(row=max_row, column=2, value=code)
            code_to_row[code] = max_row
            new_codes.append(code)

    return sheet, existing_dates, code_to_row, new_codes


def calc_score_for_index(prices, window, idx_in_valid, calc_func):
    target_idx = (window - 1) + idx_in_valid
    if target_idx >= len(prices):
        return None
    sub_prices = prices[:target_idx + 1]
    return calc_func(sub_prices, window)


# =========================
# 5. S/Z 단일 시트 저장 엔진
# =========================
def save_score_sheet(filename, dates, stocks, window, sheet_name, calc_func):
    wb = openpyxl.load_workbook(filename)

    if len(dates) < window:
        print(f"⚠ {sheet_name}: 날짜가 {window}일보다 적어 계산 불가.")
        return

    # window일 이후 날짜만 계산됨
    valid_dates = dates[window - 1:]

    sheet, existing_dates, code_to_row, new_codes = ensure_score_sheet(wb, sheet_name, stocks)
    stock_map = {str(stock["code"]): stock for stock in stocks}
    existing_count = len(existing_dates)

    # 새로 추가된 종목은 기존 열도 채워준다.
    if new_codes and existing_count > 0:
        for code in new_codes:
            stock = stock_map.get(code)
            if not stock:
                continue
            row_idx = code_to_row[code]
            prices = stock["prices"]
            for idx_global in range(existing_count):
                col_idx = 3 + idx_global
                if sheet.cell(row=row_idx, column=col_idx).value not in (None, ""):
                    continue
                score = calc_score_for_index(prices, window, idx_global, calc_func)
                if score is not None:
                    sheet.cell(row=row_idx, column=col_idx, value=score)

    if existing_count >= len(valid_dates):
        wb.save(filename)
        print(f"✅ {sheet_name}: 신규 날짜 없음 ({filename})")
        return

    for idx_global in range(existing_count, len(valid_dates)):
        date_val = valid_dates[idx_global]
        col_idx = 3 + idx_global
        header_cell = sheet.cell(row=1, column=col_idx)
        header_cell.value = date_val
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10

        for code, row_idx in code_to_row.items():
            stock = stock_map.get(code)
            if not stock:
                continue
            score = calc_score_for_index(stock["prices"], window, idx_global, calc_func)
            if score is not None:
                sheet.cell(row=row_idx, column=col_idx, value=score)

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 12

    wb.save(filename)
    print(f"✅ {sheet_name} 업데이트 완료 ({filename}, 신규 날짜 {len(valid_dates) - existing_count}개)")


# =========================
# 6. S/Z 전체 수행
# =========================
def run_total_sz(filename):
    print(f"\n=== S/Z 통합 점수 계산 시작: {filename} ===")

    dates, stocks = get_close_data(filename)

    # ---- S 점수 ----
    for window, sheet in [(20, "s20"), (60, "s60"), (120, "s120")]:
        save_score_sheet(filename, dates, stocks, window, sheet, calc_s)

    # ---- Z 점수 ----
    for window, sheet in [(20, "z20"), (60, "z60"), (120, "z120")]:
        save_score_sheet(filename, dates, stocks, window, sheet, calc_z)

    print(f"=== S/Z 통합 점수 계산 완료 ===\n")


def main():
    run_total_sz("KR_Stocks_ETF.xlsx")


if __name__ == "__main__":
    main()
